"""OCR extraction from PDF, Excel, and image documents."""

import logging
import re
import time
from pathlib import Path
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import ProcessingError

logger = logging.getLogger(__name__)

EXTRACTION_FIELDS = (
    "employee",
    "hours",
    "rate",
    "date",
    "company",
    "project",
    "ot_hours",
    "confidence",
)


class OCRService:
    def __init__(self, session: AsyncSession):
        self.session = session

    async def extract_from_file(
        self, path: str | Path, content_type: str
    ) -> dict[str, Any]:
        file_path = Path(path)
        if not file_path.exists():
            raise ProcessingError(
                f"File not found: {path}",
                reason="OCR source file does not exist on disk",
            )

        start = time.perf_counter()
        ct = content_type.lower()
        raw_text = ""
        engine = "unknown"

        try:
            if "pdf" in ct or file_path.suffix.lower() == ".pdf":
                raw_text, engine = self._extract_pdf(file_path)
            elif any(x in ct for x in ("spreadsheet", "excel")) or file_path.suffix.lower() in (
                ".xlsx",
                ".xls",
            ):
                raw_text, engine = self._extract_excel(file_path)
            elif "image" in ct or file_path.suffix.lower() in (".png", ".jpg", ".jpeg", ".tiff", ".bmp"):
                raw_text, engine = self._extract_image(file_path)
            else:
                raw_text = file_path.read_text(encoding="utf-8", errors="ignore")
                engine = "text"
        except ProcessingError:
            raise
        except Exception as exc:
            raise ProcessingError(
                f"OCR extraction failed for {file_path.name}",
                reason=str(exc),
            ) from exc

        extracted = self._parse_fields(raw_text)
        elapsed_ms = int((time.perf_counter() - start) * 1000)

        return {
            "raw_text": raw_text,
            "extracted_data": extracted,
            "ocr_engine": engine,
            "processing_time_ms": elapsed_ms,
            "confidence": extracted.get("confidence", 0.0),
        }

    def _extract_pdf(self, file_path: Path) -> tuple[str, str]:
        import pdfplumber

        parts: list[str] = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                if text.strip():
                    parts.append(text)
                tables = page.extract_tables() or []
                for table in tables:
                    for row in table:
                        parts.append(" | ".join(str(c or "") for c in row))
        return "\n".join(parts), "pdfplumber"

    def _extract_excel(self, file_path: Path) -> tuple[str, str]:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        parts: list[str] = []
        for sheet in wb.worksheets:
            parts.append(f"--- {sheet.title} ---")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    parts.append(" | ".join(cells))
        wb.close()
        return "\n".join(parts), "openpyxl"

    def _extract_image(self, file_path: Path) -> tuple[str, str]:
        try:
            import pytesseract
            from PIL import Image

            image = Image.open(file_path)
            text = pytesseract.image_to_string(image)
            return text, "pytesseract"
        except Exception as exc:
            logger.warning("pytesseract unavailable, using PIL metadata only: %s", exc)
            from PIL import Image

            image = Image.open(file_path)
            return f"Image: {image.size[0]}x{image.size[1]} {image.mode}", "pil_fallback"

    def _parse_fields(self, raw_text: str) -> dict[str, Any]:
        text = raw_text or ""
        lower = text.lower()
        fields: dict[str, Any] = {}
        confidence_factors: list[float] = []

        employee_patterns = [
            r"(?:employee|consultant|worker|name)[:\s]+([A-Za-z][A-Za-z\s\.\-']{1,60})",
            r"(?:prepared by|submitted by)[:\s]+([A-Za-z][A-Za-z\s\.\-']{1,60})",
        ]
        for pat in employee_patterns:
            m = re.search(pat, text, re.I)
            if m:
                fields["employee"] = m.group(1).strip()
                confidence_factors.append(0.85)
                break

        hours_match = re.search(
            r"(?:regular|total|billable)?\s*hours?[:\s]+(\d+(?:\.\d+)?)", lower
        )
        if hours_match:
            fields["hours"] = float(hours_match.group(1))
            confidence_factors.append(0.9)

        ot_match = re.search(
            r"(?:ot|overtime)\s*hours?[:\s]+(\d+(?:\.\d+)?)", lower
        )
        if ot_match:
            fields["ot_hours"] = float(ot_match.group(1))
            confidence_factors.append(0.88)

        rate_match = re.search(
            r"(?:hourly\s*)?rate[:\s]*\$?\s*(\d+(?:\.\d+)?)", lower
        )
        if rate_match:
            fields["rate"] = float(rate_match.group(1))
            confidence_factors.append(0.87)

        date_match = re.search(
            r"(\d{4}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/\d{2,4}|\w+\s+\d{1,2},?\s+\d{4})", text
        )
        if date_match:
            fields["date"] = date_match.group(1).strip()
            confidence_factors.append(0.82)

        company_match = re.search(
            r"(?:company|client|vendor|organization)[:\s]+([A-Za-z0-9\s&\.\-,]{2,80})",
            text,
            re.I,
        )
        if company_match:
            fields["company"] = company_match.group(1).strip().split("\n")[0]
            confidence_factors.append(0.8)

        project_match = re.search(
            r"(?:project|engagement|assignment)[:\s#]+([A-Za-z0-9\s\-_/]{2,80})",
            text,
            re.I,
        )
        if project_match:
            fields["project"] = project_match.group(1).strip().split("\n")[0]
            confidence_factors.append(0.78)

        if confidence_factors:
            fields["confidence"] = round(sum(confidence_factors) / len(confidence_factors) * 100, 2)
        elif text.strip():
            fields["confidence"] = 35.0
        else:
            fields["confidence"] = 0.0

        return fields
