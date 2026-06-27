"""SAP-style ERP export service."""

from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from uuid import UUID

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.core.exceptions import NotFoundError, ProcessingError
from app.models.invoice import ERPExport
from app.repositories.invoice_repository import ERPExportRepository, InvoiceRepository


class ERPService:
    SAP_HEADERS = [
        "Document Type",
        "Company Code",
        "Posting Date",
        "Document Date",
        "Reference",
        "Vendor Account",
        "G/L Account",
        "Amount",
        "Currency",
        "Tax Code",
        "Cost Center",
        "PO Number",
        "Line Text",
    ]

    def __init__(self, session: AsyncSession):
        self.session = session
        self.settings = get_settings()
        self.invoice_repo = InvoiceRepository(session)
        self.export_repo = ERPExportRepository(session)

    async def export_invoice(
        self,
        invoice_id: UUID,
        exported_by_id: UUID | None = None,
    ) -> ERPExport:
        invoice = await self.invoice_repo.get_with_details(invoice_id)
        if not invoice:
            raise NotFoundError("Invoice", str(invoice_id))

        rows = self._build_sap_rows(invoice)
        export_dir = Path(self.settings.local_storage_path) / "erp_exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        filename = f"SAP_{invoice.invoice_number}_{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = export_dir / filename

        try:
            self._write_sap_excel(file_path, rows)
        except Exception as exc:
            raise ProcessingError(
                "Failed to generate SAP export file",
                reason=str(exc),
            ) from exc

        mapping_config = {
            "format": "sap_excel",
            "headers": self.SAP_HEADERS,
            "company_code": "1000",
            "gl_account": "400000",
        }

        export_record = ERPExport(
            invoice_id=invoice_id,
            format="sap_excel",
            file_path=str(file_path.resolve()),
            row_count=len(rows),
            mapping_config=mapping_config,
            exported_data={"rows": rows},
            status="generated",
            exported_by_id=exported_by_id,
        )
        return await self.export_repo.create(export_record)

    async def export_invoice_response(
        self,
        invoice_id: UUID,
        exported_by_id: UUID | None = None,
        format: str = "sap_excel",
    ) -> dict:
        export = await self.export_invoice(invoice_id, exported_by_id)
        return {
            "export_id": export.id,
            "file_path": export.file_path,
            "download_url": f"/api/v1/erp/download/{export.id}",
            "row_count": export.row_count,
            "format": export.format,
        }

    def _build_sap_rows(self, invoice) -> list[dict]:
        posting_date = (
            invoice.issue_date.isoformat() if invoice.issue_date else datetime.now(UTC).date().isoformat()
        )
        line_items = invoice.line_items or []
        if not line_items:
            line_items = [
                {
                    "description": f"Services - {invoice.vendor_name or 'Vendor'}",
                    "amount": float(invoice.subtotal or invoice.total_amount or 0),
                }
            ]

        rows = []
        for idx, item in enumerate(line_items, start=1):
            amount = item.get("amount") or item.get("total") or 0
            rows.append(
                {
                    "Document Type": "KR",
                    "Company Code": "1000",
                    "Posting Date": posting_date,
                    "Document Date": posting_date,
                    "Reference": invoice.invoice_number,
                    "Vendor Account": (invoice.vendor_name or "UNKNOWN")[:10].upper(),
                    "G/L Account": "400000",
                    "Amount": float(amount),
                    "Currency": invoice.currency or "USD",
                    "Tax Code": "V1" if invoice.tax_amount else "V0",
                    "Cost Center": item.get("cost_center", "CC-100"),
                    "PO Number": invoice.po_reference or "",
                    "Line Text": item.get("description", f"Line {idx}")[:40],
                }
            )
        return rows

    def _write_sap_excel(self, file_path: Path, rows: list[dict]) -> None:
        df = pd.DataFrame(rows, columns=self.SAP_HEADERS)
        wb = Workbook()
        ws = wb.active
        ws.title = "SAP Export"

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, header in enumerate(self.SAP_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, header in enumerate(self.SAP_HEADERS, start=1):
                value = row.get(header, "")
                if header == "Amount" and value is not None:
                    value = Decimal(str(value))
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        wb.save(file_path)
